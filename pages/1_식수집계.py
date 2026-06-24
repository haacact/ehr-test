import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
import io
import json
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.header import Header
import gspread
from oauth2client.service_account import ServiceAccountCredentials

# 🚀 [추가] 엑셀 파일 디자인을 위한 openpyxl 서식 모듈
from openpyxl.styles import Alignment, Border, Side

# --- [고정 외주업체 리스트] ---
OUTSOURCING_COMPANIES = {
    "A식당(사내)": ["용성", "대명", "삼호ac", "범양", "fm", "태광", "다온정공"],
    "B식당(창녕)": ["지에쓰와이", "영진", "키움", "유나"]
}

# 🚀 식당 목록 및 식당별 담당자 이메일 주소 매핑
CAFETERIAS = ["A식당(사내)", "B식당(창녕)"]
CAFETERIA_EMAILS = {
    # 🚨 실제 식당 영양사/담당자 영문 이메일로 변경하세요
    "A식당(사내)": "un0113@hiairkorea.co.kr",
    "B식당(창녕)": "un0113@hiairkorea.co.kr"
}

# --- [구글 시트 & 이메일 연동 인프라 설정] ---
SPREADSHEET_NAME = "vacation_data"

# 사내 이메일 서버 (Gmail SMTP 기준 연동)
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
SENDER_EMAIL = "haacact@gmail.com"          
SENDER_PASSWORD = "gjurrycgnypvyilk"   

# --- [사이드바 경비 시스템 링크 추가] ---
st.sidebar.page_link("https://hiairac-expense-sysem.onrender.com/", label="경비 시스템 가기", icon="💸")
st.sidebar.divider()

@st.cache_resource
def get_gspread_client():
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    creds_dict = json.loads(st.secrets["gcp_service_account"])
    creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
    return gspread.authorize(creds)

@st.cache_data(ttl=10)
def load_meal_data():
    try:
        client = get_gspread_client()
        sheet = client.open(SPREADSHEET_NAME)
        
        ws_emp = sheet.worksheet("Employees")
        emp_data = ws_emp.get_all_values()
        df_emp = pd.DataFrame(emp_data[1:], columns=emp_data[0]) if emp_data else pd.DataFrame()
        
        ws_meals = sheet.worksheet("Meals")
        meal_data = ws_meals.get_all_values()
        df_meals = pd.DataFrame(meal_data[1:], columns=meal_data[0]) if meal_data else pd.DataFrame()
        
        for col in ['ID', 'PASSWORD', '이름', '팀', 'permission']:
            if col not in df_emp.columns: df_emp[col] = ""
            
        for col in ['ID', '이름', '팀', '날짜', '구분', '식당', '인원수', '신청시간', '상태']:
            if col not in df_meals.columns: 
                if col == '상태': df_meals[col] = "대기"
                elif col == '식당': df_meals[col] = CAFETERIAS[0]
                else: df_meals[col] = ""
            
        if not df_meals.empty:
            df_meals['인원수'] = df_meals['인원수'].replace('', '1')
            df_meals['상태'] = df_meals['상태'].replace('', '대기')
            df_meals['식당'] = df_meals['식당'].replace('', CAFETERIAS[0])
            
        return df_emp, df_meals
    except Exception as e:
        st.error(f"❌ 데이터 로드 오류: {e}")
        st.stop()

def save_meals_only(df_meals):
    try:
        client = get_gspread_client()
        sheet = client.open(SPREADSHEET_NAME)
        ws = sheet.worksheet("Meals")
        
        df_clean = df_meals.fillna("").astype(str)
        data = [df_clean.columns.values.tolist()] + df_clean.values.tolist()
        
        ws.clear()
        ws.update(values=data, value_input_option='RAW')
        load_meal_data.clear()
        return True
    except Exception as e:
        st.error(f"❌ 식수 데이터 저장 실패: {e}")
        return False

# 🚀 [공통 함수] 엑셀 시트에 디자인(너비 20, 테두리, 중앙 정렬) 입히기
def apply_excel_styling(worksheet):
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    # 셀마다 가운데 정렬 및 테두리 적용
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
    # 열 너비를 20으로 설정
    for col in worksheet.columns:
        worksheet.column_dimensions[col[0].column_letter].width = 20

def send_cafeteria_order_email(target_date, cafe_name, total_count, team_counts_df):
    target_email = CAFETERIA_EMAILS.get(cafe_name, SENDER_EMAIL)
    subject = f"[식수발주] {target_date} 하이에어공조 중식 발주 요청 ({cafe_name})"
    
    body = f"안녕하세요. {cafe_name} 담당자님,\n하이에어공조(주) 금일({target_date}) 중식 최종 발주 인원을 안내해 드립니다.\n\n"
    body += f"■ 총 발주 인원 : {total_count} 명\n\n"
    body += "세부 부서별/외주업체별 명단은 첨부된 엑셀 파일을 확인해 주시기 바랍니다.\n\n"
    body += "맛있고 안전한 식사 준비 부탁드리겠습니다.\n감사합니다.\n\n- 하이에어공조(주) 총괄 관리자 드림 -"

    try:
        msg = MIMEMultipart()
        msg['From'] = SENDER_EMAIL
        msg['To'] = target_email
        msg['Subject'] = Header(subject, 'utf-8')
        msg.attach(MIMEText(body, 'plain', 'utf-8'))
        
        if not team_counts_df.empty:
            excel_buffer = io.BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                team_counts_df.to_excel(writer, sheet_name='당일중식_발주현황', index=False)
                # 🚀 작성된 시트에 디자인 함수 호출
                apply_excel_styling(writer.sheets['당일중식_발주현황'])
            
            excel_buffer.seek(0)
            attachment = MIMEApplication(excel_buffer.read(), _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            
            cafe_code = "Cafe_A" if "A식당" in cafe_name else "Cafe_B"
            file_name = f"Lunch_Order_{target_date}_{cafe_code}.xlsx"
            
            attachment.add_header('Content-Disposition', 'attachment', filename=file_name)
            msg.attach(attachment)

        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(SENDER_EMAIL, SENDER_PASSWORD)
        server.sendmail(SENDER_EMAIL, target_email, msg.as_string())
        server.quit()
        return True
    except Exception as e:
        st.error(f"중식 메일 발송 오류: {e}")
        return False

def send_cafeteria_dinner_email(date_range_str, cafe_name, daily_totals, pivot_df):
    target_email = CAFETERIA_EMAILS.get(cafe_name, SENDER_EMAIL)
    subject = f"[식수발주] 하이에어공조 차주 석식({date_range_str}) 발주 요청 ({cafe_name})"
    
    body = f"안녕하세요. {cafe_name} 담당자님,\n하이에어공조(주) 차주({date_range_str}) 석식 요일별 최종 발주 인원을 안내해 드립니다.\n\n"
    body += "■ 요일별 총 발주 인원 요약\n"
    for day_str, count in daily_totals.items():
        body += f"- {day_str} : {count}명\n"
        
    body += "\n자세한 요일별/부서별 석식 신청 매트릭스(Matrix)는 첨부된 엑셀 파일을 확인해 주시기 바랍니다.\n\n"
    body += "맛있고 안전한 식사 준비 부탁드리겠습니다.\n감사합니다.\n\n- 하이에어공조(주) 총괄 관리자 드림 -"

    try:
        msg = MIMEMultipart()
        msg['From'] = SENDER_EMAIL
        msg['To'] = target_email
        msg['Subject'] = Header(subject, 'utf-8')
        msg.attach(MIMEText(body, 'plain', 'utf-8'))
        
        if not pivot_df.empty:
            excel_buffer = io.BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                summary_df = pd.DataFrame(list(daily_totals.items()), columns=['요일', '발주인원(명)'])
                summary_df.to_excel(writer, sheet_name='1_Daily_Summary', index=False)
                pivot_df.to_excel(writer, sheet_name='2_Dinner_Matrix', index=False)
                
                # 🚀 두 개의 시트 모두에 디자인 함수 호출
                apply_excel_styling(writer.sheets['1_Daily_Summary'])
                apply_excel_styling(writer.sheets['2_Dinner_Matrix'])
                
            excel_buffer.seek(0)
            attachment = MIMEApplication(excel_buffer.read(), _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            
            cafe_code = "Cafe_A" if "A식당" in cafe_name else "Cafe_B"
            safe_date_range = date_range_str.replace("~", "_").replace("/", "")
            file_name = f"Dinner_Order_{safe_date_range}_{cafe_code}.xlsx"
            
            attachment.add_header('Content-Disposition', 'attachment', filename=file_name)
            msg.attach(attachment)

        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(SENDER_EMAIL, SENDER_PASSWORD)
        server.sendmail(SENDER_EMAIL, target_email, msg.as_string())
        server.quit()
        return True
    except Exception as e:
        st.error(f"석식 메일 발송 오류: {e}")
        return False

df_emp, df_meals = load_meal_data()

if 'logged_in' not in st.session_state:
    st.session_state['logged_in'] = False
    
if 'user_info' not in st.session_state:
    st.session_state['user_info'] = None

if not st.session_state['logged_in']:
    st.title("🍱 사내 식수 관리 시스템")
    st.warning("💡 연차 시스템 메인화면에서 로그인하시면 편리하게 이용하실 수 있습니다.")
    with st.form("meal_login"):
        i_id = st.text_input("ID(사번)")
        i_pw = st.text_input("비밀번호", type="password")
        if st.form_submit_button("로그인"):
            i_pw_stripped = i_pw.lstrip('0') if i_pw.lstrip('0') != '' else i_pw
            user = df_emp[(df_emp['ID'] == i_id) & ((df_emp['PASSWORD'] == i_pw) | (df_emp['PASSWORD'] == i_pw_stripped))]
            
            if not user.empty:
                st.session_state.update({'logged_in': True, 'user_info': user.iloc[0]})
                st.rerun()
            else: st.error("정보가 올바르지 않습니다.")
    st.stop()

user_info = df_emp[df_emp['ID'] == st.session_state['user_info']['ID']].iloc[0]

st.sidebar.title(f"👤 {user_info['이름']} ({user_info['permission']})")
if st.sidebar.button("로그아웃"):
    st.session_state['logged_in'] = False
    st.rerun()

st.title("🍱 사내 식수 신청 / 집계 시스템")
st.caption(f"접속자: {user_info['이름']}님 ({user_info['팀']} / {user_info['permission']})")

menu = ["✍️ 식수 신청/취소"]
if user_info['permission'] == "관리자":
    menu += ["📊 식수 인원 집계 현황"]
    
choice = st.radio("기능 선택", menu, horizontal=True)
st.divider()

KST = ZoneInfo("Asia/Seoul")
now = datetime.now(KST)
today_date = now.date()
is_past_930 = now.hour > 9 or (now.hour == 9 and now.minute >= 30)

st.write("현재시간", now)
st.write("오늘", today_date)
st.write("09:30 이후?", is_past_930)

this_mon = today_date - timedelta(days=today_date.weekday())
this_fri = this_mon + timedelta(days=4)
next_mon = this_mon + timedelta(days=7)
next_fri = next_mon + timedelta(days=4)

if choice == "✍️ 식수 신청/취소":
    tabs_names = ["☀️ 팀 중식 (당일)", "🌙 팀 석식 (차주)"]
    
    is_pac_team = '생산(PAC외)' in user_info['팀'] or '생산(퓨얼셀)' in user_info['팀'] or user_info['permission'] == '관리자'
    
    if is_pac_team:
        tabs_names.append("🚚 외주업체 (생산 전용)")
        
    tabs = st.tabs(tabs_names)
    
    with tabs[0]:
        c_left, c_right = st.columns([1.1, 1.2])
        
        with c_left:
            st.subheader("📝 우리 팀 중식 일괄 신청")
            select_date = st.date_input("신청 날짜 선택", value=today_date, key="lunch_date_picker")
            date_str = select_date.strftime("%Y-%m-%d")
            s_cafe_lunch = st.radio("🍽️ 이용할 식당 선택", CAFETERIAS, horizontal=True, key="lunch_cafe")
            
            if select_date == today_date and is_past_930:
                st.error("🚨 금일 식수 집계가 마감되었습니다. (09:30 마감)")
            elif select_date < today_date:
                st.error("🚨 지나간 날짜의 식수는 신청할 수 없습니다.")
            elif select_date.weekday() >= 5:
                st.error("📅 주말은 중식을 신청할 수 없습니다.")
            else:
                my_team_members = df_emp[df_emp['팀'] == user_info['팀']].copy()
                if my_team_members.empty:
                    st.info("조회된 팀원이 없습니다.")
                else:
                    applied_ids = df_meals[(df_meals['날짜'] == date_str) & (df_meals['구분'] == '중식')]['ID'].astype(str).tolist()
                    my_team_members['선택'] = False
                    my_team_members['신청상태'] = my_team_members['ID'].apply(lambda x: "🚨 이미 신청됨" if str(x) in applied_ids else "가능")
                    
                    edited_df = st.data_editor(
                        my_team_members[['선택', 'ID', '이름', '신청상태']],
                        column_config={"선택": st.column_config.CheckboxColumn(), "ID": "사번", "이름": "이름", "신청상태": "상태"},
                        disabled=["ID", "이름", "신청상태"], hide_index=True, use_container_width=True, key=f"lunch_ed_{date_str}"
                    )
                    
                    if st.button(f"🍱 선택한 팀원 {s_cafe_lunch} 제출", use_container_width=True):
                        selected_rows = edited_df[(edited_df['선택'] == True) & (edited_df['신청상태'] == "가능")]
                        if selected_rows.empty:
                            st.warning("⚠️ 신청할 팀원이 없습니다.")
                        else:
                            new_rows = []
                            for _, r in selected_rows.iterrows():
                                new_rows.append({
                                    "ID": str(r['ID']), "이름": r['이름'], "팀": user_info['팀'],
                                    "날짜": date_str, "구분": "중식", "식당": s_cafe_lunch, "인원수": "1",
                                    "신청시간": now.strftime("%Y-%m-%d %H:%M:%S"), "상태": "대기"
                                })
                            df_meals = pd.concat([df_meals, pd.DataFrame(new_rows)], ignore_index=True)
                            if save_meals_only(df_meals):
                                st.success(f"{s_cafe_lunch} 중식 신청 완료!")
                                st.rerun()

        with c_right:
            st.subheader("🔍 팀 중식 신청 내역 확인")
            l_date_range = st.date_input("📅 조회 기간 (시작일 - 종료일)", value=(this_mon, this_fri), key="l_date_filter")
            
            if isinstance(l_date_range, (tuple, list)):
                if len(l_date_range) == 2: s_d, e_d = l_date_range
                elif len(l_date_range) == 1: s_d = e_d = l_date_range[0]
                else: s_d = e_d = today_date
            else:
                s_d = e_d = l_date_range
            
            if user_info['permission'] == "관리자":
                team_list = ["✨ 전사 전체보기"] + sorted(df_emp['팀'].unique().tolist())
                selected_team = st.selectbox("🗂️ 조회할 부서 선택 (관리자)", team_list, key="l_team_sel")
            else:
                selected_team = user_info['팀']
                st.info(f"🚩 **{user_info['팀']}**의 중식 내역이 표시됩니다.")

            view_meals = df_meals[
                (df_meals['구분'] == '중식') &
                (~df_meals['이름'].str.contains(r'\[외주\]', na=False)) &
                (df_meals['날짜'] >= s_d.strftime('%Y-%m-%d')) &
                (df_meals['날짜'] <= e_d.strftime('%Y-%m-%d'))
            ]
            
            if user_info['permission'] == "관리자" and selected_team != "✨ 전사 전체보기":
                view_meals = view_meals[view_meals['팀'] == selected_team]
            elif user_info['permission'] != "관리자":
                view_meals = view_meals[view_meals['팀'] == user_info['팀']]
                
            view_meals = view_meals.sort_values(by="날짜", ascending=False)
            
            if view_meals.empty:
                st.info("조건에 맞는 중식 신청 내역이 없습니다.")
            else:
                with st.container(height=450):
                    for idx, row in view_meals.iterrows():
                        target_date = datetime.strptime(row['날짜'], "%Y-%m-%d").date()
                        is_locked = row.get('상태', '대기') == '마감' or target_date < today_date or (target_date == today_date and is_past_930)
                            
                        c_cols = st.columns([3, 2, 2.5, 1.5])
                        c_cols[0].write(f"📅 {row['날짜']}")
                        c_cols[1].write(f"👤 {row['이름']}")
                        c_cols[2].write(f"🍽️ {row.get('식당', 'A식당(사내)')}")
                        
                        if not is_locked:
                            if c_cols[3].button("❌ 취소", key=f"del_l_{idx}"):
                                df_meals = df_meals.drop(idx)
                                if save_meals_only(df_meals): st.rerun()
                        else:
                            if row.get('상태') == '마감': c_cols[3].error("🔒 마감")
                            else: c_cols[3].warning("🔒 시간초과")

    with tabs[1]:
        c_left, c_right = st.columns([1.1, 1.2])
        
        with c_left:
            st.subheader("📝 차주 석식 인원 일괄 입력")
            st.info(f"📅 다음 주 **{next_mon.strftime('%m/%d')} (월) ~ {(next_mon + timedelta(days=4)).strftime('%m/%d')} (금)** 석식 인원")
            s_cafe_dinner = st.radio("🍽️ 이용할 식당 선택", CAFETERIAS, horizontal=True, key="dinner_cafe")
            
            days_name = ["월요일", "화요일", "수요일", "목요일", "금요일"]
            dinner_counts = {}
            next_week_strs = [ (next_mon + timedelta(days=i)).strftime("%Y-%m-%d") for i in range(5) ]
            
            locked_dinner_days = df_meals[(df_meals['팀'] == user_info['팀']) & (df_meals['날짜'].isin(next_week_strs)) & (df_meals['구분'] == '석식') & (df_meals['상태'] == '마감')]
            
            with st.form("dinner_form"):
                for i, day_name in enumerate(days_name):
                    target_date = next_mon + timedelta(days=i)
                    t_date_str = target_date.strftime("%Y-%m-%d")
                    
                    existing_row = df_meals[(df_meals['팀'] == user_info['팀']) & (df_meals['날짜'] == t_date_str) & (df_meals['구분'] == '석식') & (df_meals.get('식당', '') == s_cafe_dinner) & (~df_meals['이름'].str.contains(r'\[외주\]', na=False))]
                    default_val = 0
                    if not existing_row.empty:
                        try: default_val = int(float(existing_row.iloc[0]['인원수']))
                        except: default_val = 0
                        
                    dinner_counts[t_date_str] = st.number_input(f"▶️ {target_date.strftime('%m/%d')} ({day_name}) 석식 인원(명)", min_value=0, value=default_val, step=1)
                
                if st.form_submit_button(f"🌙 차주 {s_cafe_dinner} 석식 일괄 제출 / 수정", use_container_width=True):
                    if not locked_dinner_days.empty:
                        st.error("🚨 관리자가 이미 마감 확정한 날짜가 포함되어 있어 일괄 수정이 불가합니다.")
                    else:
                        df_meals = df_meals[~((df_meals['팀'] == user_info['팀']) & (df_meals['날짜'].isin(next_week_strs)) & (df_meals['구분'] == '석식') & (df_meals.get('식당', '') == s_cafe_dinner) & (~df_meals['이름'].str.contains(r'\[외주\]', na=False)))]
                        
                        new_dinner_rows = []
                        for d_str, count in dinner_counts.items():
                            if count > 0:
                                new_dinner_rows.append({
                                    "ID": user_info['ID'], "이름": user_info['이름'], "팀": user_info['팀'],
                                    "날짜": d_str, "구분": "석식", "식당": s_cafe_dinner, "인원수": str(count),
                                    "신청시간": now.strftime("%Y-%m-%d %H:%M:%S"), "상태": "대기"
                                })
                        if new_dinner_rows:
                            df_meals = pd.concat([df_meals, pd.DataFrame(new_dinner_rows)], ignore_index=True)
                        if save_meals_only(df_meals):
                            st.components.v1.html("<script>alert('🌙 차주 석식 인원 등록이 완료되었습니다.');</script>", height=0, width=0)
                            st.rerun()

        with c_right:
            st.subheader("🔍 팀 석식 신청 내역 확인")
            d_date_range = st.date_input("📅 조회 기간 (시작일 - 종료일)", value=(next_mon, next_fri), key="d_date_filter")
            
            if isinstance(d_date_range, (tuple, list)):
                if len(d_date_range) == 2: s_d, e_d = d_date_range
                elif len(d_date_range) == 1: s_d = e_d = d_date_range[0]
                else: s_d = e_d = today_date
            else:
                s_d = e_d = d_date_range
            
            if user_info['permission'] == "관리자":
                team_list = ["✨ 전사 전체보기"] + sorted(df_emp['팀'].unique().tolist())
                selected_team = st.selectbox("🗂️ 조회할 부서 선택 (관리자)", team_list, key="d_team_sel")
            else:
                selected_team = user_info['팀']
                st.info(f"🚩 **{user_info['팀']}**의 석식 내역이 표시됩니다.")

            view_meals = df_meals[
                (df_meals['구분'] == '석식') &
                (~df_meals['이름'].str.contains(r'\[외주\]', na=False)) &
                (df_meals['날짜'] >= s_d.strftime('%Y-%m-%d')) &
                (df_meals['날짜'] <= e_d.strftime('%Y-%m-%d'))
            ]
            
            if user_info['permission'] == "관리자" and selected_team != "✨ 전사 전체보기":
                view_meals = view_meals[view_meals['팀'] == selected_team]
            elif user_info['permission'] != "관리자":
                view_meals = view_meals[view_meals['팀'] == user_info['팀']]
                
            view_meals = view_meals.sort_values(by="날짜", ascending=False)
            
            if view_meals.empty:
                st.info("조건에 맞는 석식 신청 내역이 없습니다.")
            else:
                with st.container(height=450):
                    for idx, row in view_meals.iterrows():
                        target_date = datetime.strptime(row['날짜'], "%Y-%m-%d").date()
                        is_locked = row.get('상태', '대기') == '마감' or target_date < today_date or (target_date == today_date and is_past_930)
                            
                        c_cols = st.columns([2.5, 2.5, 2.0, 1.5, 1.5])
                        c_cols[0].write(f"📅 {row['날짜']}")
                        c_cols[1].write(f"🏢 {row['팀']}")
                        c_cols[2].write(f"🍽️ {row.get('식당', 'A식당(사내)')}")
                        
                        if not is_locked:
                            try: current_val = int(float(row['인원수']))
                            except: current_val = 1
                            new_count = c_cols[3].number_input("명", min_value=0, value=current_val, key=f"d_edit_count_{idx}", label_visibility="collapsed")
                            if c_cols[4].button("수정", key=f"btn_d_edit_{idx}"):
                                df_meals.at[idx, '인원수'] = str(new_count)
                                if save_meals_only(df_meals): st.rerun()
                        else:
                            c_cols[3].write(f"🔢 {row['인원수']}명")
                            if row.get('상태') == '마감': c_cols[4].error("🔒 마감")
                            else: c_cols[4].warning("🔒 초과")

    if is_pac_team:
        with tabs[2]:
            c_left, c_right = st.columns([1.1, 1.2])
            
            with c_left:
                st.subheader("🚚 고정 외주업체 식수 일괄 등록")
                o_type = st.radio("식수 구분 선택", ["☀️ 당일 중식", "🌙 차주 석식 (월~금 일괄)"], horizontal=True)
                s_cafe_out = st.radio("🍽️ 외주업체 이용 식당", CAFETERIAS, horizontal=True, key="out_cafe")
                
                current_companies = OUTSOURCING_COMPANIES.get(s_cafe_out, [])
                
                if o_type == "☀️ 당일 중식":
                    o_date = st.date_input("신청 날짜", value=today_date, key="o_date")
                    if o_date == today_date and is_past_930:
                        st.error("🚨 금일 식수 집계가 마감되었습니다.")
                    elif o_date < today_date:
                        st.error("🚨 지나간 날짜입니다.")
                    elif o_date.weekday() >= 5:
                        st.error("📅 주말은 식수를 신청할 수 없습니다.")
                    else:
                        with st.form("o_lunch_form"):
                            st.write(f"▼ {s_cafe_out} 업체별 중식 인원(명) 입력")
                            grid_cols = st.columns(4)
                            o_lunch_counts = {}
                            
                            for i, comp in enumerate(current_companies):
                                existing = df_meals[(df_meals['날짜'] == o_date.strftime("%Y-%m-%d")) & (df_meals['구분'] == '중식') & (df_meals.get('식당', '') == s_cafe_out) & (df_meals['이름'] == f"[외주] {comp}")]
                                default_val = int(float(existing.iloc[0]['인원수'])) if not existing.empty else 0
                                o_lunch_counts[comp] = grid_cols[i % 4].number_input(comp, min_value=0, value=default_val, step=1, key=f"ol_{comp}")
                                
                            if st.form_submit_button(f"🍱 외주업체 중식({s_cafe_out}) 일괄 등록", type="primary", use_container_width=True):
                                target_names = [f"[외주] {c}" for c in current_companies]
                                df_meals = df_meals[~((df_meals['날짜'] == o_date.strftime("%Y-%m-%d")) & (df_meals['구분'] == '중식') & (df_meals.get('식당', '') == s_cafe_out) & (df_meals['이름'].isin(target_names)))]
                                
                                new_rows = []
                                for comp, count in o_lunch_counts.items():
                                    if count > 0:
                                        new_rows.append({
                                            "ID": user_info['ID'], "이름": f"[외주] {comp}", "팀": user_info['팀'],
                                            "날짜": o_date.strftime("%Y-%m-%d"), "구분": "중식", "식당": s_cafe_out, "인원수": str(count),
                                            "신청시간": now.strftime("%Y-%m-%d %H:%M:%S"), "상태": "대기"
                                        })
                                if new_rows:
                                    df_meals = pd.concat([df_meals, pd.DataFrame(new_rows)], ignore_index=True)
                                if save_meals_only(df_meals):
                                    st.success("외주업체 중식 등록 완료!")
                                    st.rerun()
                                    
                else: 
                    st.write(f"📅 다음 주 **{next_mon.strftime('%m/%d')} (월) ~ {(next_mon + timedelta(days=4)).strftime('%m/%d')} (금)** 석식")
                    
                    next_week_strs = [ (next_mon + timedelta(days=i)).strftime("%Y-%m-%d") for i in range(5) ]
                    locked_dinner_days = df_meals[(df_meals['이름'].str.contains(r'\[외주\]', na=False)) & (df_meals['날짜'].isin(next_week_strs)) & (df_meals['구분'] == '석식') & (df_meals['상태'] == '마감')]
                    
                    if not locked_dinner_days.empty:
                        st.error("🚨 관리자가 이미 마감 확정한 날짜가 포함되어 일괄 수정이 불가합니다.")
                    else:
                        days_name = ["월", "화", "수", "목", "금"]
                        col_names = [f"{next_week_strs[i][-5:]} ({days_name[i]})" for i in range(5)]
                        
                        init_data = []
                        for comp in current_companies:
                            row_data = {"업체명": comp}
                            for i, d_str in enumerate(next_week_strs):
                                existing = df_meals[(df_meals['날짜'] == d_str) & (df_meals['구분'] == '석식') & (df_meals.get('식당', '') == s_cafe_out) & (df_meals['이름'] == f"[외주] {comp}")]
                                val = int(float(existing.iloc[0]['인원수'])) if not existing.empty else 0
                                row_data[col_names[i]] = val
                            init_data.append(row_data)
                            
                        df_o_dinner = pd.DataFrame(init_data)
                        
                        st.write("▼ 아래 표의 숫자를 클릭하여 요일별 인원을 바로 입력하세요.")
                        edited_o_dinner = st.data_editor(df_o_dinner, hide_index=True, disabled=["업체명"], use_container_width=True)
                        
                        if st.button(f"🌙 외주업체 차주 석식({s_cafe_out}) 일괄 저장", type="primary", use_container_width=True):
                            target_names = [f"[외주] {c}" for c in current_companies]
                            df_meals = df_meals[~((df_meals['날짜'].isin(next_week_strs)) & (df_meals['구분'] == '석식') & (df_meals.get('식당', '') == s_cafe_out) & (df_meals['이름'].isin(target_names)))]
                            
                            new_rows = []
                            for idx, row in edited_o_dinner.iterrows():
                                comp = row['업체명']
                                for i, d_str in enumerate(next_week_strs):
                                    count = row[col_names[i]]
                                    if count > 0:
                                        new_rows.append({
                                            "ID": user_info['ID'], "이름": f"[외주] {comp}", "팀": user_info['팀'],
                                            "날짜": d_str, "구분": "석식", "식당": s_cafe_out, "인원수": str(count),
                                            "신청시간": now.strftime("%Y-%m-%d %H:%M:%S"), "상태": "대기"
                                        })
                            if new_rows:
                                df_meals = pd.concat([df_meals, pd.DataFrame(new_rows)], ignore_index=True)
                            if save_meals_only(df_meals):
                                st.success("외주업체 석식 등록 완료!")
                                st.rerun()

            with c_right:
                st.subheader("🔍 외주업체 식수 신청 내역")
                o_date_range = st.date_input("📅 조회 기간 (시작일 - 종료일)", value=(this_mon, next_fri), key="o_date_filter")
                
                if isinstance(o_date_range, (tuple, list)):
                    if len(o_date_range) == 2: s_d, e_d = o_date_range
                    elif len(o_date_range) == 1: s_d = e_d = o_date_range[0]
                    else: s_d = e_d = today_date
                else:
                    s_d = e_d = o_date_range
                
                st.info("🚩 등록된 **외주업체** 식수 내역만 표시됩니다.")

                view_meals = df_meals[
                    (df_meals['이름'].str.contains(r'\[외주\]', na=False)) &
                    (df_meals['날짜'] >= s_d.strftime('%Y-%m-%d')) &
                    (df_meals['날짜'] <= e_d.strftime('%Y-%m-%d'))
                ]
                    
                view_meals = view_meals.sort_values(by=["날짜", "구분"], ascending=[False, True])
                
                if view_meals.empty:
                    st.info("조건에 맞는 외주업체 식수 내역이 없습니다.")
                else:
                    with st.container(height=450):
                        for idx, row in view_meals.iterrows():
                            target_date = datetime.strptime(row['날짜'], "%Y-%m-%d").date()
                            is_locked = row.get('상태', '대기') == '마감' or target_date < today_date or (target_date == today_date and is_past_930)
                                
                            c_cols = st.columns([2.5, 2.0, 2.0, 1.5, 1.5])
                            c_cols[0].write(f"📅 {row['날짜']} [{row['구분']}]")
                            c_cols[1].write(f"🏢 {row['이름']}")
                            c_cols[2].write(f"🍽️ {row.get('식당', 'A식당(사내)')}")
                            
                            if not is_locked:
                                try: current_val = int(float(row['인원수']))
                                except: current_val = 1
                                new_count = c_cols[3].number_input("명", min_value=1, value=current_val, key=f"o_edit_count_{idx}", label_visibility="collapsed")
                                if c_cols[4].button("수정", key=f"btn_o_edit_{idx}"):
                                    df_meals.at[idx, '인원수'] = str(new_count)
                                    if save_meals_only(df_meals): st.rerun()
                            else:
                                c_cols[3].write(f"🔢 {row['인원수']}명")
                                if row.get('상태') == '마감': c_cols[4].error("🔒 마감")
                                else: c_cols[4].warning("🔒 초과")

# --- 📊 식수 인원 집계 현황 탭 (관리자 전용) ---
elif choice == "📊 식수 인원 집계 현황":
    st.subheader("📊 관리자 마스터 식수 집계 현황")
    st.info("💡 식당별 분리 집계 및 개별 식당 담당자 발송 기능이 탑재되어 있습니다.")
    
    tab_lunch_dash, tab_dinner_dash, tab_excel_dash = st.tabs([
        "☀️ 당일 중식 집계 (일자별/식당별)", 
        "🌙 차주 석식 집계 (식당별 Matrix)", 
        "📥 누적 데이터 백업"
    ])
    
    with tab_lunch_dash:
        col_f1, col_f2 = st.columns(2)
        search_date = col_f1.date_input("조회 날짜", value=today_date, key="lunch_dash_picker")
        dash_cafe_lunch = col_f2.selectbox("조회 식당", ["전체보기"] + CAFETERIAS, key="dash_cafe_lunch")
        s_date_str = search_date.strftime("%Y-%m-%d")
        
        lunch_meals = df_meals[(df_meals['날짜'] == s_date_str) & (df_meals['구분'] == '중식')].copy()
        if dash_cafe_lunch != "전체보기":
            lunch_meals = lunch_meals[lunch_meals['식당'] == dash_cafe_lunch]
            
        lunch_meals['인원수_num'] = pd.to_numeric(lunch_meals['인원수'], errors='coerce').fillna(1).astype(int)
        total_lunch = lunch_meals['인원수_num'].sum()
        
        st.markdown(f"#### 📅 {s_date_str} [{dash_cafe_lunch}] 중식 신청 총 인원: `{total_lunch}명`")
        
        if lunch_meals.empty:
            st.info("해당 조건의 중식 신청 내역이 없습니다.")
        else:
            lunch_meals['집계그룹'] = lunch_meals.apply(lambda x: x['이름'] if '[외주]' in x['이름'] else x['팀'], axis=1)
            team_counts = lunch_meals.groupby(['식당', '집계그룹'])['인원수_num'].sum().reset_index()
            team_counts.columns = ['식당', '부서/외주업체명', '신청 인원(명)']
            
            col_btn1, col_btn2 = st.columns(2)
            with col_btn1:
                if st.button(f"🔒 {s_date_str} 중식 집계 마감 (전체 수정 불가)", type="primary", use_container_width=True):
                    target_idx = df_meals[(df_meals['날짜'] == s_date_str) & (df_meals['구분'] == '중식')].index
                    df_meals.loc[target_idx, '상태'] = '마감'
                    if save_meals_only(df_meals):
                        st.success("✅ 해당 날짜의 중식 집계가 마감되었습니다.")
                        st.rerun()
            with col_btn2:
                if st.button(f"📧 각 식당 담당자에게 [중식] 발주 메일 일괄 전송 (엑셀첨부)", use_container_width=True):
                    success_cnt = 0
                    for c_name in CAFETERIA_EMAILS.keys():
                        c_df = team_counts[team_counts['식당'] == c_name]
                        c_total = c_df['신청 인원(명)'].sum()
                        if c_total > 0:
                            if send_cafeteria_order_email(s_date_str, c_name, c_total, c_df):
                                success_cnt += 1
                    if success_cnt > 0:
                        st.success(f"🎉 총 {success_cnt}곳의 식당에 중식 발주 메일(엑셀 포함) 전송 완료!")
                    else:
                        st.error("🚨 발송 대상 데이터가 없거나 전송에 실패했습니다.")
                    
            col_l1, col_l2 = st.columns([1, 1.2])
            with col_l1:
                st.markdown("##### 🚩 부서/외주업체별 집계")
                st.dataframe(team_counts, use_container_width=True, hide_index=True)
                
            with col_l2:
                st.markdown("##### 👥 상세 명단")
                st.dataframe(lunch_meals[['식당', '집계그룹', '이름', '인원수_num']].rename(columns={'집계그룹':'소속', '인원수_num':'인원수'}), use_container_width=True, hide_index=True)
                
    with tab_dinner_dash:
        next_mon = today_date + timedelta(days=(7 - today_date.weekday()))
        next_week_dates = [(next_mon + timedelta(days=i)).strftime("%Y-%m-%d") for i in range(5)]
        next_week_days = ["월", "화", "수", "목", "금"]
        
        st.markdown(f"#### 📅 차주 석식 부서/업체별 일괄 집계판")
        st.caption(f"기준 범위: {next_mon.strftime('%Y-%m-%d')} ~ {(next_mon + timedelta(days=4)).strftime('%Y-%m-%d')}")
        
        dash_cafe_dinner = st.selectbox("조회 식당", ["전체보기"] + CAFETERIAS, key="dash_cafe_dinner")
        
        dinner_next_week = df_meals[(df_meals['구분'] == '석식') & (df_meals['날짜'].isin(next_week_dates))].copy()
        if dash_cafe_dinner != "전체보기":
            dinner_next_week = dinner_next_week[dinner_next_week['식당'] == dash_cafe_dinner]
        
        if dinner_next_week.empty:
            st.info("해당 조건의 차주 석식 내역이 없습니다.")
        else:
            dinner_next_week['인원수_num'] = pd.to_numeric(dinner_next_week['인원수'], errors='coerce').fillna(0).astype(int)
            
            daily_totals = {}
            for i, d_str in enumerate(next_week_dates):
                daily_totals[f"{d_str[-5:].replace('-', '/')} ({next_week_days[i]})"] = dinner_next_week[dinner_next_week['날짜'] == d_str]['인원수_num'].sum()
                
            dinner_next_week['집계그룹'] = dinner_next_week.apply(lambda x: x['이름'] if '[외주]' in x['이름'] else x['팀'], axis=1)
            
            pivot_dinner = dinner_next_week.pivot_table(index=['식당', '집계그룹'], columns='날짜', values='인원수_num', aggfunc='sum').fillna(0).astype(int)
            for d_str in next_week_dates:
                if d_str not in pivot_dinner.columns: pivot_dinner[d_str] = 0
            pivot_dinner = pivot_dinner[next_week_dates]
            
            new_cols = []
            for i, d_str in enumerate(next_week_dates):
                _, m, d = d_str.split('-')
                new_cols.append(f"{m}/{d} ({next_week_days[i]})")
            pivot_dinner.columns = new_cols
            pivot_dinner = pivot_dinner.reset_index().rename(columns={'집계그룹': '부서/외주업체명'})

            col_btn1, col_btn2 = st.columns(2)
            with col_btn1:
                if st.button(f"🔒 차주(5일) 석식 전체 마감 확정", type="primary", use_container_width=True):
                    target_idx = df_meals[(df_meals['구분'] == '석식') & (df_meals['날짜'].isin(next_week_dates))].index
                    df_meals.loc[target_idx, '상태'] = '마감'
                    if save_meals_only(df_meals):
                        st.success("✅ 차주 석식 집계가 전면 마감되었습니다.")
                        st.rerun()
            with col_btn2:
                if st.button(f"📧 각 식당 담당자에게 [차주 석식] 발주 메일 일괄 전송 (엑셀첨부)", use_container_width=True):
                    success_cnt = 0
                    date_range_str = f"{next_mon.strftime('%m/%d')}~{(next_mon + timedelta(days=4)).strftime('%m/%d')}"
                    
                    for c_name in CAFETERIA_EMAILS.keys():
                        c_df = pivot_dinner[pivot_dinner['식당'] == c_name].drop(columns=['식당'], errors='ignore')
                        
                        c_daily_totals = {}
                        for i, d_str in enumerate(next_week_dates):
                            val = dinner_next_week[(dinner_next_week['날짜'] == d_str) & (dinner_next_week['식당'] == c_name)]['인원수_num'].sum()
                            c_daily_totals[f"{d_str[-5:].replace('-', '/')} ({next_week_days[i]})"] = val
                        
                        if sum(c_daily_totals.values()) > 0:
                            if send_cafeteria_dinner_email(date_range_str, c_name, c_daily_totals, c_df):
                                success_cnt += 1
                                
                    if success_cnt > 0:
                        st.success(f"🎉 총 {success_cnt}곳의 식당에 석식 발주 메일(엑셀 포함) 전송 완료!")
                    else:
                        st.warning("발송할 식수 데이터가 없습니다.")
            
            st.markdown("##### 🏢 차주 석식 종합 Matrix")
            st.dataframe(pivot_dinner, use_container_width=True, hide_index=True)
            
            st.divider()
            st.markdown(f"##### 📦 [{dash_cafe_dinner}] 요일별 최종 발주 총 인원")
            d_cols = st.columns(5)
            for i, d_str in enumerate(next_week_dates):
                total_day_dinner = dinner_next_week[dinner_next_week['날짜'] == d_str]['인원수_num'].sum()
                d_cols[i].metric(f"{next_week_days[i]}요일", f"{total_day_dinner} 명")
            
    with tab_excel_dash:
        st.markdown("##### 🗓️ 전체 데이터 백업")
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df_meals.to_excel(writer, sheet_name="Meals", index=False)
        buffer.seek(0)
        st.download_button("📥 누적 식수 데이터 엑셀 다운로드", data=buffer, file_name=f"meal_data_backup_{datetime.now().strftime('%Y%m%d')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
